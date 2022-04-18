import base64
import logging
import os, shutil
import re
import pandas as pd
import numpy as np
import cv2
import loging
from OCR import *
import PIL.Image as tu
from openpyxl.drawing.image import Image
import traceback
import openpyxl
import web_check

log = loging.log()

PATHDIR = os.path.join(os.path.expanduser("~"), 'Desktop') + "\\screenshot\\"

if os.path.exists(PATHDIR):
    shutil.rmtree(PATHDIR)

TABLIST =   {
        "净含量":"D6",
        "地址":"D8",
        "生产者":"D9",
        "联系方式":"D10",
        "生产日期":"D11",
        "贮存条件":"D13",
        "食品生产许可证编号":"D15",
        "产品标准代号":"D17",
        "辐照食品":"D19",
        "转基因食品":"D21",
        "质量（品质）等级":"D26"
              }


class Img_ocr:
    errlist = []
    def __init__(self,img_path):
        (
            self.path,
            self.file_name
        )= os.path.split(img_path)
        with open(img_path, 'rb') as f:
            base64_data = base64.b64encode(f.read())
            res = bytes.decode(base64_data)
        self.img_result = general_ocr_v3(res) 
        if 'details' in self.img_result:
            self.img_dir = PATHDIR+self.file_name.split('.')[0]
            os.makedirs(self.img_dir)
            self.details = self.img_result['details']
            self.description = self.img_result['description']
            self.img = cv2.imdecode(np.fromfile(img_path, dtype=np.uint8),1)
            self.exp =  False
        else:self.exp =  True

    def re2(self, re_list:list):
        result = []
        a = str(self.details).replace("'","").replace(" ","")
        [[result.append(re.findall('\w.*',i)[0][0:-1]) for i in re.findall(r"%s(.*?),"%t,a) if len(re.findall('\w+',i)) > 0] for t in re_list if t]
        return result

    def productList(self):
        b = {}
        ls = pd.read_excel("标签描述字典.xlsx").to_dict('records')
        for i in ls:
            if not (i['标签描述'] is np.nan):
                b[i["标签名称"]] = self.re2(i['标签描述'].split(","))
     
        b["食品生产许可证编号"] =  [sc[0].upper() for i in self.details if (sc := re.findall(r"SC\d{14}",i[1],re.I))]
        b["产品标准代号"] =  [gb[0].upper() for i in self.details if (gb := re.findall(r"[A-Z]{1}B[/T]{0,2}\s{0,1}\d{5}",i[1],re.I))]
        # b["食品生产许可证编号"] =  re.findall(r"SC\d{14}",self.description)
        # b["产品标准代号"] =  re.findall(r"[A-Z]{1}B[/T]{0,2}\s{0,1}\d{5}",self.description)
        return b

    
    def __xy(self,b):
        for _ in self.details:
            if b in _[1]:
                return _[0]
        return False
    
    def main(self):
        if self.exp:
            self.errlist.append(self.file_name)
            log.warning("图片可能超过4M 跳过执行")
            return 0
        product = self.productList()
        for i in product:
            if product[i]:
                log.info(i+" 结果: "+product[i][0])
                xy = self.__xy(product[i][0])
                if xy == False:
                    continue
                img = self.img[int(xy[0][1]-xy[1][1]/2):int(xy[0][1]+xy[1][1]/2),int(xy[0][0]-xy[1][0]/2):int(xy[0][0]+xy[1][0]/2)]
                cv2.imencode('.jpg',img)[1].tofile(os.path.join(self.img_dir,i+".jpg"))
                log.info(f"{i}   标签截取成功")
        return product



class __excel:
    def __init__(self,imgdir, product, description) -> None:
        self.wd = openpyxl.load_workbook('食品表格.xlsx')
        self.ws = self.wd.active
        self.imgdir = imgdir
        self.product = product
        self.description = description

    def in_pic(self):
        for i in os.listdir(self.imgdir):
            cell = TABLIST[i.split(".")[0]]
            filename = self.imgdir +"\\"+ i
            imgSize = tu.open(filename).size
            img = Image(filename)
            newsize = (imgSize[0]*(150/imgSize[0]), imgSize[1]*(150/imgSize[0]))
            if newsize[1] <30:
                newsize =(150, 30)
            log.info(i + "  size修改为" + str(imgSize)+"=>" + str(newsize))
            self.ws.column_dimensions["D"].width = 150/8   
            self.ws.row_dimensions[int(cell[1:])].height = newsize[1]*0.75        
            img.width, img.height = newsize            
            self.ws.add_image(img, cell)

    def check(self):
        new_form = {k:"E" + v[1:] for k, v in TABLIST.items()}
        ws = self.ws
        description = self.description
        productdict = self.product
        unit = ["生产者","地址","联系方式"]
        log.info(productdict)
        log.info(productdict[unit[0]] + productdict[unit[1]] +productdict[unit[2]])
        if productdict[unit[0]] and productdict[unit[1]] and productdict[unit[2]]:
            ws[new_form[unit[1]]] = "符合"
        else:
            ws[new_form[unit[1]]] = "不符合"

        unit = "净含量"
        if productdict[unit] and len(re.findall(r"\d+",productdict[unit][0])[0]) <= 3:
            ws[new_form[unit]] = "符合"
        else:
            ws[new_form[unit]] = "不符合"
            
        unit = "生产日期"  
        if productdict[unit] :
            ws[new_form[unit]] = "符合"
        else:
            ws[new_form[unit]] = "不符合"
            
        unit = "贮存条件"  
        if productdict[unit] :
            ws[new_form[unit]] = "符合"
        else:
            ws[new_form[unit]] = "不符合"
            
        unit = "食品生产许可证编号"  
        if productdict[unit] :
            df = web_check.query_licence(productdict[unit][0])
            log.warning(df)
            log.warning(productdict["地址"])
            log.warning(df["生产地址"])

            if df and [i for i in productdict["地址"] if i in df["生产地址"]]:
                ws[new_form[unit]] = "符合"
            else:
                ws[new_form[unit]] = "不符合"


        unit = "产品标准代号"  
        if productdict[unit] :
            code  = productdict[unit]
            log.warning(code[0][:-5] +"+"+code[0][-5:])
            fs = web_check.food_standards(code[0][:-5] +"+"+code[0][-5:])
            log.info(fs)
            if code and fs:
                ws[new_form[unit]] = "符合"
            else:
                ws[new_form[unit]] = "不符合"
            

        unit = "辐照食品"  
        if unit in description:
            if productdict[unit] :
                ws[new_form[unit]] = "符合"
            else:
                ws[new_form[unit]] = "不符合"
            

        unit = "转基因食品"  
        if unit in description:
            if productdict[unit] :
                ws[new_form[unit]] = "符合"
            else:
                ws[new_form[unit]] = "不符合"
            
            
        unit = "质量（品质）等级"
        if unit in description:
            if productdict[unit]:
                ws[new_form[unit]] = "符合"
            else:
                ws[new_form[unit]] = "不符合"
        return True

    def main(self):
        self.in_pic()
        self.check()
        self.wd.save(self.imgdir+"/审核结果.xlsx")
        self.wd.close()

if __name__ == '__main__':

    dir = 'Input/'
    for img_path in os.listdir(dir):
        if img_path.split(".")[-1] in ['jpg','png']:
            log.info(f"~~~~~~~~~~~任务< {img_path} >开始~~~~~~~~~~~")
            ll = Img_ocr(dir+img_path)
            product = ll.main()
            if product:
                __excel(ll.img_dir,product,ll.description).main()
    log.error("未处理图片: "+ ",  ".join(ll.errlist))
    logging.shutdown()
