from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from lxml import etree
import cv2
import numpy as np
# from matplotlib import pyplot as plt
# import scipy as sp
import pandas as pd
import os
# from matplotlib import image as ig
import re
import base64
import shutil
import json
import requests
from openpyxl.drawing.image import Image
import PIL.Image as tu
import openpyxl
import PySimpleGUI as sg
import traceback

HEADER = {
    'Content-Type': 'application/json',
    'X-HW-ID': 'com.huawei.gts.experience4ca',
    'X-HW-APPKEY': '3LYg4LYTPfCgYGA1DePUrAnkryRrTLQqpNLhKNU0DVl35xR4uCsxSDv8GX9JajV1'
}


biao =   {
        "净含量":"D6",
        "地址":"D8",
        "生产者":"D9",
        "联系方式":"D10",
        "生产日期":"D11",
        "贮存条件":"D13",
        "食品生产许可证编号":"D15",
        "产品标准代号":"D17",
        "辐照食品":"D19",
        "转基因食品":"D21",
        "质量（品质）等级":"D26"
              }


pattern = 'windows'
'headless'

def general_ocr_v3(image_base64):
    "https://apigw.huawei.com/api/v2/ocr/general-text"
    url = 'https://apigw-02.huawei.com/api/ocr-dev/api/v1/ocr/general-text'
    body = {
        "image_base64": image_base64,
        "mode": "noah-doc"
    }

    resp = requests.post(url,
                         headers=HEADER,
                         data=json.dumps(body),
                         verify=False)
    t = json.loads(resp.text)
    print(t)
    resp.close()
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    return t


def table_detect(image_base64):
    url = 'https://apigw-02.huawei.com/api/ocr-dev/v1/ocr/table-detection'
    body = {
        "image_base64": image_base64,
        "correct_orientation": "true",
        "clean_stamp": "false",
        "perspective_correct": "false",
        "bending_correct": "false",
    }

    resp = requests.post(url,
                         headers=HEADER,
                         data=json.dumps(body),
                         verify=False)
    t = json.loads(resp.text)
    print(t)
    resp.close()
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    return t

def GetDesktopPath():
    return os.path.join(os.path.expanduser("~"), 'Desktop')


def opencv_3(img, a, spath,b):
    path = get_path(img,"i")
    print(path)
    os.makedirs(path[0])
    img1 = cv2.imdecode(np.fromfile(img, dtype=np.uint8),1)
    # print(img1)
    product = productList(a,b)
    for i in product:
        if product[i]:
            print(i,product[i][0])
            xy = baidu_xy(product[i][0],a)
            if xy == False:
                continue
            img2 = img1[xy[0]:xy[1],xy[2]:xy[3]]
            cv2.imencode('.jpg',img2)[1].tofile(get_path(img,i)[1])
            print("标签截取成功")
    wd = openpyxl.load_workbook(spath)
    w = wd.active
    # ipath = ipath(r"C:\Users\\screenshot\111")
    # print(ipath[1])
    imglist = os.listdir(path[0])
    print(imglist)
    for i in imglist:
        xlw(path[0]+i, biao[i.split(".")[0]], w)
    for _ in range(3):
        try:
            if shbq(product,b,w):
                break
        except Exception as e:
            traceback.print_exc()
            print("正在重试")
    wd.save(path[0]+"审核结果.xlsx")
    wd.close()
    return product


def baidu_xy(b,a):
    for i in a:
        # print(b,_[1])
        if b in i['words']:
            t = [i['location']['top'],i['location']['top']+i['location']['height'], i['location']['left'],i['location']['left']+i['location']['width']]
            return [i['location']['top'],i['location']['top']+i['location']['height'], i['location']['left'],i['location']['left']+i['location']['width']]
    return False

def opencv_2(img, a, spath,b):
    path = get_path(img,"i")
    print(path)
    os.makedirs(path[0])
    img1 = cv2.imdecode(np.fromfile(img, dtype=np.uint8),1)
    # print(img1)
    product = productList(a,b)
    for i in product:
        if product[i]:
            print(i,product[i][0])
            xy = _xy(product[i][0],a)
            if xy == False:
                continue
            img2 = img1[int(xy[0][1]-xy[1][1]/2):int(xy[0][1]+xy[1][1]/2),int(xy[0][0]-xy[1][0]/2):int(xy[0][0]+xy[1][0]/2)]
            cv2.imencode('.jpg',img2)[1].tofile(get_path(img,i)[1])
            print("标签截取成功")
    wd = openpyxl.load_workbook(spath)
    w = wd.active
    # ipath = ipath(r"C:\Users\\screenshot\111")
    # print(ipath[1])
    imglist = os.listdir(path[0])
    print(imglist)
    for i in imglist:
        xlw(path[0]+i, biao[i.split(".")[0]], w)
    for _ in range(3):
        try:
            if shbq(product,b,w):
                break
        except Exception as e:
            traceback.print_exc()
            print("正在重试")
    wd.save(path[0]+"审核结果.xlsx")
    wd.close()
    return product

def re2(a,b = list):
    c = []
    a = str(a).replace("'","").replace(" ","")
    [[c.append(re.findall('\w.*',i)[0][0:-1]) for i in re.findall(r"%s(.*?),"%t,a) if len(re.findall('\w+',i)) > 0] for t in b]
    return c

# def re2(a,b = list): #修改计划暂不使用
    # for i in b:
    #     for y in a:
    #         p = re.findall(i,y[1])
    #         if p:
    #             c.append(y[1])
    # print(c)
#     [[[363.0, 964.0000610351562], [597.7855834960938, 34.00593566894531], -0.38388559222221375], '【地址】深圳市龙岗区宝龙街道宝龙社区锦龙一路9号多利工业厂区C栋2-3楼']

def productList(a,c):
    b = {}
    b["净含量"] = re2(a, ["净.量"])
    b["地址"] =  re2(a, ["地址"])
    b["生产者"] = re2(a,["生产商", "被委托方", '被委托生产者', "被委托生产企业", "厂家"])
    b["联系方式"] =  re2(a, ["联系方式","客服热线","电话"])
    b["生产日期"] =  re2(a, ["生产日期"])
    b["贮存条件"] =  re2(a, ["贮.条件","贮.方法"])
    b["食品生产许可证编号"] =  re.findall(r"SC\d{14}",c)
    b["产品标准代号"] =  re.findall(r"[A-Z]{1}B[/T]{0,2}\s{0,1}\d{5}",c)
    b["辐照食品"] =  re2(a, ["辐照食品"])
    b["转基因食品"] =  re2(a, ["转基因食品"])
    b["质量（品质）等级"] =  re2(a, ["质量（品质）等级"])
    return b


def _xy(b,a):
    for _ in a:
        # print(b,_[1])
        if b in _[1]:
            # print(_)
            return _[0]
    return False


def xlw(a,b,w):
    filename = a
    imgSize = tu.open(filename).size
    # print(imgSize)
    img = Image(filename)
    newsize = (imgSize[0]*(150/imgSize[0]), imgSize[1]*(150/imgSize[0]))
    if newsize[1] <30:
        newsize =(150, 30)
#     wd = openpyxl.load_workbook(r".\input\食品表格.xlsx")
#     w = wd.active
    print(newsize)
    w.column_dimensions["D"].width = 150/8   
    w.row_dimensions[int(b[1:])].height = newsize[1]*0.75        
    img.width, img.height = newsize            
    w.add_image(img, b)


def ipath(ipath1):
    a = GetDesktopPath()+"\\screenshot\\" + ipath1.split("\\")[-1].split(".")[0] +'.xlsx'
    b = GetDesktopPath()+"\\screenshot\\" + ipath1.split("\\")[-1]
    return a,b
    
def get_path(a,name):
    print(a,"~~~",name)
    b = a.split("//")[-1].split(".")[0]
    print(b)
    c = str(r'%s'%(os.path.join(os.path.expanduser("~"), 'Desktop') + f"\\screenshot\\{b}\\"))
    d = c +f"{name}.jpg"
    return c,d

def sol(Label_catalog = list, spath = str):
    p = {}
    if os.path.exists(str(r'%s'%(os.path.join(os.path.expanduser("~"), 'Desktop') + "\\screenshot\\"))):
        shutil.rmtree(str(r'%s'%(os.path.join(os.path.expanduser("~"), 'Desktop') + "\\screenshot\\")))
    for Label in Label_catalog:
        if Label.split(".")[-1] in ["jpg","png"]:
            print(Label,"~~~~~~~~~~~新的开始")
            with open(Label, 'rb') as fileObj:
                image_data = fileObj.read()
                base64_data = base64.b64encode(image_data)
                # print(bytes.decode(base64_data))
            res = bytes.decode(base64_data)
            text_ = general_ocr_v3(res)
            if "details" in text_:
                # print(text_['description'])
                p[Label.split("\\")[-1].split(".")[0]] = opencv_2(Label,text_["details"],spath,text_['description'])
    # for t in p.values():
    #     print(t)
            print(Label,"~~~~~~~~~~~完成")
    return p


def shbq(a,b,w):
    new_form = {k:"E" + v[1:] for k,v in biao.items()}
    # print(a)
    # wd = openpyxl.load_workbook(r".\input\食品表格.xlsx")
    # w = wd.active
    unit = ["生产者","地址","联系方式"]
    print(a[unit[0]] , a[unit[1]] , a[unit[2]])
    print(new_form[unit[1]])
    if a[unit[0]] and a[unit[1]] and a[unit[2]]:
        w[new_form[unit[1]]] = "符合"
    else:
        w[new_form[unit[1]]] = "不符合"

    unit = "净含量"
    if a[unit] and len(re.findall(r"\d+",a[unit][0])[0]) <= 3:
        w[new_form[unit]] = "符合"
    else:
        w[new_form[unit]] = "不符合"
        
    unit = "生产日期"  
    if a[unit] :
        w[new_form[unit]] = "符合"
    else:
        w[new_form[unit]] = "不符合"
        
    unit = "贮存条件"  
    if a[unit] :
        w[new_form[unit]] = "符合"
    else:
        w[new_form[unit]] = "不符合"
        
    unit = "食品生产许可证编号"  
    re3 = re.findall("SC\d{14}",a[unit][0])
    if a[unit] and re3:
        sh = vlicence(a[unit][0])
        if sh and [i for i in a["地址"] if i in sh]:
            w[new_form[unit]] = "符合"
        else:
            w[new_form[unit]] = "不符合"


    unit = "产品标准代号"  
    if a[unit] :
        # print(a[unit])
        code  = a[unit]
        print(code)
        if code and Food_standards(code[0][:-5] +"+"+code[0][-5:]):
            w[new_form[unit]] = "符合"
        else:
            w[new_form[unit]] = "不符合"
        

    unit = "辐照食品"  
    if unit in b:
        if a[unit] :
            w[new_form[unit]] = "符合"
        else:
            w[new_form[unit]] = "不符合"
        

    unit = "转基因食品"  
    if unit in b:
        if a[unit] :
            w[new_form[unit]] = "符合"
        else:
            w[new_form[unit]] = "不符合"
        
        
    unit = "质量（品质）等级"  
    if unit in b:
        if a[unit] :
            w[new_form[unit]] = "符合"
        else:
            w[new_form[unit]] = "不符合"
    return True


def vlicence(name):
#chromedriver的路径
    print(name)
    # try:
    url = 'http://spscxkcx.gsxt.gov.cn/spscxk/detail.xhtml?zsbh=%s&lb=SC'%name
    # chromedriver = r"C:\chromedriver\chromedriver.exe"
    # os.environ["webdriver.chrome.driver"] = chromedriver
    chrome_options = Options()
    chrome_options.add_argument(pattern)
    driver = webdriver.Chrome(chrome_options=chrome_options)
    # driver.maximize_window()
#控制浏览器写入并转到链接
    driver.get(url)
    r = driver.page_source
    print("打开%s页面成功"%url)
    html = etree.HTML(r)
    result = html.xpath('/html/body/div[2]/div/table/tbody/tr[5]/td/text()')
    driver.close()
    if result:
        print(result,"~~~~~~~~~~~~生产许可证查询地址")
        return result[0]
    else:
        return ""
# except:
    # return False


def Food_standards(name):
    print(name)
#chromedriver的路径
    url = 'http://down.foodmate.net/standard/search.php?corpstandard=2&fields=0&kw=%s'%name
    print("打开%s页面"%url)
    # chromedriver = r"C:\chromedriver\chromedriver.exe"
    # os.environ["webdriver.chrome.driver"] = chromedriver
    chrome_options = Options()
    chrome_options.add_argument(pattern)
    driver = webdriver.Chrome(chrome_options=chrome_options)
    driver.maximize_window()
#控制浏览器写入并转到链接
    driver.get(url)
    r = driver.page_source
    html = etree.HTML(r)
    result = html.xpath('//*[@class="bgt"]')
    print(len(result))
    driver.close()
    if len(result) > 0 :
        return True
    else:
        return False
    

def mian(a):
    files = sg.popup_get_folder("请选择标签图片文件夹")
    imglist = [files + "//" + i for i in os.listdir(files)]
    print(imglist)
    sol(Label_catalog = imglist, spath = a)

if __name__=="__main__":
    # a = [['80mm', [2274, 117]], ['134mm', [3217, 118]], ['暖新来。', [3214, 462]], ['WARM', [3218, 494]], ['不添加任何防腐剂', [2817, 612]], ['100年前它在，100年后它还在', [3173, 1049]], ['【产品名称】', [2794, 1133]], ['暖起来初心姜母', [2960, 1134]], ['224mm', [3900, 1156]], ['', [2748, 1169]], ['地】 1广东东莞', [2896, 1171]], ['【配', [2752, 1205]], ['料】 小黄姜红糖', [2888, 1207]], ['浮起来', [1350, 1214]], ['【保 质期】 18个月', [2843, 1241]], ['WARM', [1372, 1277]], ['【生产日期】 见袋体侧边', [2871, 1278]], ['【贮藏方法】 本品应存放在干燥、阴凉处。', [2955, 1315]], ['UP', [1335, 1321]], ['【出品商】 深圳市暖起来食品科技有限公司', [2979, 1351]], ['（地址', [2800, 1388]], ['深圳市福田区福田街道岗厦社区彩田路3069号星河世纪大厦Α栋3009F', [3263, 1387]], ['生产商】', [2803, 1422]], ['CHINA SYSTEM', [1366, 1424]], ['乘胜（东莞）食品科技有限公司', [3037, 1422]], ['【地 址】 东莞市望牛墩镇望牛墩北环路27号3号楼101', [3051, 1457]], ['中国制', [1366, 1465]], ['【产品标准号】 Q/CSDG 0001S', [2902, 1493]], ['【食品生产许可证编号】 SC10744190014749', [2997, 1529]], ['【食用方式】', [2789, 1566]], ['开袋后冲调食用，建议冲入200m⌊沸水冲调，亦可根据个人口感调整。', [3266, 1566]], ['初心姜母可与茶、牛奶、鸡蛋、甜酒、面包、螃蟹等随意搭配食用。', [3256, 1602]], ['初|心|姜|母', [1373, 1612]], ['THE MATERNAL GINGER OF FIRET MOOD', [1376, 1674]], ['客服热线：400-663-9994 www.nuanqilai.net', [3403, 1703]], ['净含量：300g(30gx10袋)', [1374, 1724]], ['项目 【营养成分表】 NRV%', [2916, 1719]], ['每100克（g）', [2921, 1741]], ['"能量', [2795, 1767]], ['967干焦（kJ）', [2920, 1769]], ['12%', [3028, 1769]], ['固形物含量不低于50%', [1376, 1783]], ['蛋白质', [2796, 1797]], ['0.6克(g)', [2918, 1798]], ['1%', [3027, 1796]], ['脂肪', [2795, 1825]], ['0克（g）', [2921, 1826]], ['0%', [3025, 1824]], ['56.3克（g）', [2921, 1854]], ['碳水化合物', [2796, 1857]], ['19%', [3027, 1853]], ['6973819825550', [3294, 1868]], ['19毫克（mg） 1%', [2956, 1884]]]
    # drawing(a, r"C:\Users\war7n\Desktop\me.png")
    # imglist = os.listdir(r"C:\Users\war7n\OneDrive\weAutomate\LableAudit\Input")
    # b = r"C:\Users\war7n\OneDrive\weAutomate\LableAudit\Input"
    # print(imglist)
    # for i in imglist:
    #     img_path = b+"\\"+i

    #     if i.split(".")[-1] in ["jpg","png"]:
    #         with open(img_path, 'rb') as fileObj:
    #             image_data = fileObj.read()
    #             base64_data = base64.b64encode(image_data)
    #             # print(bytes.decode(base64_data))
    #         res = bytes.decode(base64_data)
    #         text_ = general_ocr_v3(res)
    #         # print(text_["details"])
    #         if "details" in text_:
    #             opencv_2(img_path,text_["details"])
            # for t in text_["details"]:
            #     print(t)
    # a = ['C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\111.png', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\222.jpg', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\333.jpg', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\444.jpg', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\555.jpg', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\666.jpg', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\777.png', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\888.jpg', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\999.png', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\GB 7718-2011食品标签审核原始记录表.docx', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\me.png', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\parameters.json', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\标签111.jpg', 'C:\\Users\\war7n\\OneDrive\\weAutomate\\标签初审\\input\\食品表格.xlsx']
    b = r"C:\Users\war7n\OneDrive\Working files\weAutomate\标签初审\Input\食品表格.xlsx"
    files = sg.popup_get_folder("请选择标签图片文件夹")
    imglist = [files + "//" + i for i in os.listdir(files)]
    print(imglist)
    sol(Label_catalog = imglist, spath = b)

